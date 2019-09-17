import urllib.request
import urllib.parse
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

wb = openpyxl.load_workbook('./list1.xlsx')
sheet = wb['Sheet1']

baseUrl = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query='
companyrow = 1 #없어도 될 것 같은데

for cr in sheet: #얘도 마찬가지로 없어도 될듯
    companyrow += 1
    
    if companyrow <= 1257: #1257이 최종
        plusUrl = sheet.cell(row=companyrow, column=1).value
        url = baseUrl + urllib.parse.quote_plus(plusUrl)
        html = urllib.request.urlopen(url).read()
        soup = BeautifulSoup(html, 'html.parser')
        addr = soup.find_all(class_='txt_ellipsis')
        print('='*50)
        print(sheet.cell(row=companyrow, column=1).value)
        
        for i in addr:
            try:
                print(i.attrs['title'])
                sheet.cell(row=companyrow, column=3, value=i.attrs['title'])
            except:
                pass
            
    else:
        print('='*50)
        print('마지막 입니다.')

wb.save('./result.xlsx')
    

#판다스로 엑셀 데이터 불러오기
#data = pd.read_excel('./list.xlsx')
#print(data)
#companyname = data[]
#print(companyname)

baseUrl = 'https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query='
plusUrl = '삼성전자'
#plusUrl = input('검색어를 입력하세요: ')
url = baseUrl + urllib.parse.quote_plus(plusUrl)

html = urllib.request.urlopen(url).read()
soup = BeautifulSoup(html, 'html.parser')

title = soup.find_all(class_='txt_ellipsis')

for i in title:
    print(i.attrs['title'])
    print()

# 리스트의 기본 구조
addr = [i.attrs['title']]
addr.append('주소')
print(addr)